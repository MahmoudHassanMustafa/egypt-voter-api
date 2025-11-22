#!/usr/bin/env python3
"""
FREE Electoral Data Scraper using Selenium (No AI/API costs!)
Extracts electoral information from elections.eg website and populates Excel file.
100% Free - No API keys or AI services required!
"""

import time
import sys
from pathlib import Path
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
import logging
import argparse
import re

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - [%(name)s] %(message)s'
)
logger = logging.getLogger(__name__)

# Allowed districts/sections to process
ALLOWED_DISTRICTS = [
    "قسم الشرق",
    "قسم العرب",
    "قسم الضواحى",
    "قسم أول بورفؤاد",
    "قسم ثان بورفؤاد"
]

# Website URL
INQUIRY_URL = "https://www.elections.eg/inquiry"


class FreeElectionsScraper:
    """Free scraper using Selenium - no AI/API costs!"""
    
    def __init__(self, headless: bool = False, max_retries: int = 3, retry_delay: int = 2):
        """
        Initialize the scraper with browser options.
        
        Args:
            headless: Run browser in headless mode (no GUI)
            max_retries: Maximum number of retry attempts for failed requests
            retry_delay: Base delay in seconds between retries (uses exponential backoff)
        """
        self.driver = None
        self.max_retries = max_retries
        self.retry_delay = retry_delay
        self.setup_driver(headless)
    
    def setup_driver(self, headless: bool = False):
        """Setup Chrome WebDriver with appropriate options."""
        chrome_options = Options()
        if headless:
            chrome_options.add_argument('--headless=new')  # Use new headless mode

        # Essential options for running Chrome in Docker containers
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--disable-software-rasterizer')
        chrome_options.add_argument('--remote-debugging-port=9222')
        chrome_options.add_argument('--disable-background-timer-throttling')
        chrome_options.add_argument('--disable-backgrounding-occluded-windows')
        chrome_options.add_argument('--disable-renderer-backgrounding')
        chrome_options.add_argument('--disable-features=TranslateUI')
        chrome_options.add_argument('--disable-ipc-flooding-protection')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_argument('--disable-extensions')
        chrome_options.add_argument('--disable-plugins')
        chrome_options.add_argument('--disable-images')
        chrome_options.add_argument('--disable-plugins-discovery')
        chrome_options.add_argument('--disable-print-preview')
        chrome_options.add_argument('--disable-component-extensions-with-background-pages')
        chrome_options.add_argument('--no-first-run')
        chrome_options.add_argument('--disable-default-apps')
        chrome_options.add_argument('--disable-sync')
        chrome_options.add_argument('--hide-crash-restore-bubble')

        # Memory and performance optimizations
        chrome_options.add_argument('--memory-pressure-off')
        chrome_options.add_argument('--max_old_space_size=4096')
        chrome_options.add_argument('--disable-web-security')
        chrome_options.add_argument('--allow-running-insecure-content')
        chrome_options.add_argument('--disable-background-networking')
        chrome_options.add_argument('--disable-client-side-phishing-detection')
        chrome_options.add_argument('--disable-component-update')

        # User agent to avoid detection
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')

        # Set user data directory to a writable location
        chrome_options.add_argument('--user-data-dir=/tmp/chrome-user-data')
        chrome_options.add_argument('--data-path=/tmp/chrome-data')
        chrome_options.add_argument('--disk-cache-dir=/tmp/chrome-cache')
        chrome_options.add_argument('--remote-debugging-address=0.0.0.0')
        
        try:
            # Try to use webdriver-manager first (automatic ChromeDriver management)
            try:
                from webdriver_manager.chrome import ChromeDriverManager
                from selenium.webdriver.chrome.service import Service
                import os

                # Set custom cache path for webdriver-manager to work with read-only filesystem
                cache_path = "/app/.wdm"

                # Set environment variables for webdriver-manager
                os.environ["WDM_LOCAL"] = "1"  # Use local cache
                os.environ["WDM_CACHE_DIR"] = cache_path  # Set cache directory

                # Also set HOME to /app so webdriver-manager uses /app/.wdm by default
                os.environ["HOME"] = "/app"

                # Ensure the directory exists
                os.makedirs(cache_path, exist_ok=True)

                # Try with custom path parameter (if supported)
                try:
                    driver_path = ChromeDriverManager(path=cache_path).install()
                except TypeError:
                    # Fallback if path parameter is not supported
                    driver_path = ChromeDriverManager().install()

                # Ensure the chromedriver executable has proper permissions
                import stat
                if driver_path:
                    # Make sure the chromedriver binary is executable
                    os.chmod(driver_path, stat.S_IRWXU | stat.S_IRGRP | stat.S_IXGRP | stat.S_IROTH | stat.S_IXOTH)

                service = Service(driver_path, log_path='/tmp/chromedriver.log', log_level='DEBUG')
                self.driver = webdriver.Chrome(service=service, options=chrome_options)
                logger.info("Chrome WebDriver initialized using webdriver-manager")
            except ImportError:
                # Fallback to system ChromeDriver
                self.driver = webdriver.Chrome(options=chrome_options)
                logger.info("Chrome WebDriver initialized using system ChromeDriver")
            
            self.driver.maximize_window()
            logger.info("Chrome WebDriver initialized successfully")
        except Exception as e:
            logger.error(f"Failed to initialize WebDriver: {e}")
            logger.error("Make sure ChromeDriver is installed and in PATH, or install webdriver-manager")
            raise
    
    def scrape_electoral_data(self, national_id: str) -> dict:
        """
        Scrape electoral data for a given national ID with retry mechanism.
        
        Args:
            national_id: The national ID to query
            
        Returns:
            Dictionary with extracted electoral information
        """
        last_error = None
        
        for attempt in range(self.max_retries):
            try:
                if attempt > 0:
                    # Exponential backoff: wait longer with each retry
                    delay = self.retry_delay * (2 ** (attempt - 1))
                    logger.info(f"Retry attempt {attempt + 1}/{self.max_retries} after {delay}s delay")
                    time.sleep(delay)
                
                logger.info(f"Querying national ID: {national_id} (attempt {attempt + 1}/{self.max_retries})")
                
                # Navigate to the inquiry page
                try:
                    self.driver.get(INQUIRY_URL)
                    time.sleep(2)  # Wait for page to load
                except Exception as e:
                    logger.warning(f"Failed to navigate to page: {e}")
                    last_error = f"Navigation error: {str(e)}"
                    continue
                
                wait = WebDriverWait(self.driver, 15)
                
                # Switch to the iframe (the form is inside an iframe)
                try:
                    iframe = wait.until(EC.presence_of_element_located((By.ID, "ocv_iframe_id")))
                    self.driver.switch_to.frame(iframe)
                    logger.info("Switched to iframe")
                except TimeoutException:
                    logger.warning("Could not find iframe - retrying")
                    last_error = "Could not find iframe - page structure may have changed"
                    try:
                        self.driver.switch_to.default_content()
                    except:
                        pass
                    continue
                
                # Find and fill the national ID input field
                try:
                    nid_input = wait.until(EC.presence_of_element_located((By.ID, "nid")))
                    nid_input.clear()
                    nid_input.send_keys(national_id)
                    logger.info(f"Entered national ID: {national_id}")
                except TimeoutException:
                    logger.warning("Could not find national ID input field - retrying")
                    last_error = "Could not find national ID input field"
                    try:
                        self.driver.switch_to.default_content()
                    except:
                        pass
                    continue
                
                # Find and click the submit button
                try:
                    submit_btn = wait.until(EC.element_to_be_clickable((By.ID, "submit_btn")))
                    submit_btn.click()
                    logger.info("Clicked submit button")
                except TimeoutException:
                    logger.warning("Could not find submit button - retrying")
                    last_error = "Could not find submit button"
                    try:
                        self.driver.switch_to.default_content()
                    except:
                        pass
                    continue
                
                # Wait for results to load
                time.sleep(3)
                
                # Wait a bit more for dynamic content
                try:
                    wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'مركزك الإنتخابي') or contains(text(), 'قسم') or contains(text(), 'عفوا') or contains(text(), 'الرقم القومي')]")))
                except:
                    pass  # Continue even if timeout
                
                # Extract data from results
                result_data = self.extract_result_data()
                
                # Switch back to default content
                try:
                    self.driver.switch_to.default_content()
                except:
                    pass
                
                # Check if extraction was successful
                if result_data.get('success'):
                    logger.info(f"Successfully retrieved data for {national_id}")
                    return result_data
                else:
                    # Extraction failed, but don't retry if it's a known error state
                    error = result_data.get('error', '')
                    if 'parse' in error.lower() or 'extract' in error.lower():
                        # These are likely data extraction issues, retry might help
                        last_error = error
                        logger.warning(f"Data extraction failed: {error} - retrying")
                        continue
                    else:
                        # Other errors, return immediately
                        return result_data
                
            except Exception as e:
                logger.error(f"Error scraping data for {national_id} (attempt {attempt + 1}): {str(e)}")
                last_error = str(e)
                try:
                    self.driver.switch_to.default_content()
                except:
                    pass
                
                # If it's the last attempt, don't continue
                if attempt == self.max_retries - 1:
                    break
        
        # All retries exhausted
        logger.error(f"Failed to scrape data for {national_id} after {self.max_retries} attempts")
        return {
            'success': False,
            'error': f'Failed after {self.max_retries} attempts. Last error: {last_error}',
            'national_id': national_id,
            'retries_exhausted': True
        }
    
    def extract_result_data(self) -> dict:
        """
        Extract data from the results page.
        
        Returns:
            Dictionary with extracted fields
        """
        try:
            # Get page source to check for error messages first
            try:
                page_text = self.driver.find_element(By.TAG_NAME, "body").text
            except Exception as e:
                logger.error(f"Failed to get page text: {e}")
                return {
                    'success': False,
                    'error': 'Failed to read page content'
                }
            
            # Check if page is empty or has minimal content
            if not page_text or len(page_text.strip()) < 10:
                logger.warning("Page appears to be empty or incomplete")
                return {
                    'success': False,
                    'error': 'Page content is empty or incomplete'
                }
            
            # Check for error messages
            if "عفوا, غير مسموح لإقل من 18 سنة بالإنتخاب" in page_text:
                logger.info("Found underage error message")
                return {
                    'success': True,
                    'data': {
                        'status': 'underage',
                        'error_message': 'عفوا, غير مسموح لإقل من 18 سنة بالإنتخاب'
                    }
                }
            
            if "الرقم القومي غير مدرج بقاعدة بيانات الناخبين" in page_text:
                logger.info("Found not registered error message")
                return {
                    'success': True,
                    'data': {
                        'status': 'not_registered',
                        'error_message': 'الرقم القومي غير مدرج بقاعدة بيانات الناخبين'
                    }
                }
            
            # Extract electoral information
            # The data is displayed in a structured format
            data = {
                'status': 'success',
                'electoral_center': '',
                'district': '',
                'address': '',
                'subcommittee_number': '',
                'electoral_list_number': ''
            }
            
            # Get full page text
            full_text = page_text
            
            # Method 1: Try to extract using XPath (more reliable)
            try:
                # Look for elements that contain the labels and their values
                # The structure is usually: "Label: Value" on the same line or nearby
                
                # Extract electoral center (مركزك الإنتخابي)
                try:
                    elem = self.driver.find_element(By.XPATH, "//*[contains(text(), 'مركزك الإنتخابي')]")
                    # Get parent or next sibling that contains the value
                    parent_text = elem.find_element(By.XPATH, "./..").text
                    if ':' in parent_text:
                        parts = parent_text.split(':', 1)
                        if len(parts) > 1:
                            data['electoral_center'] = parts[1].strip().split('\n')[0].strip()
                except:
                    # Fallback to regex
                    match = re.search(r'مركزك الإنتخابي[:\s]+([^\n]+)', full_text)
                    if match:
                        data['electoral_center'] = match.group(1).strip()
                
                # Extract district (قسم)
                try:
                    elem = self.driver.find_element(By.XPATH, "//*[contains(text(), 'قسم:')]")
                    parent_text = elem.find_element(By.XPATH, "./..").text
                    if 'قسم:' in parent_text:
                        parts = parent_text.split('قسم:', 1)
                        if len(parts) > 1:
                            data['district'] = parts[1].strip().split('\n')[0].strip()
                except:
                    match = re.search(r'قسم[:\s]+([^\n]+)', full_text)
                    if match:
                        data['district'] = match.group(1).strip()
                
                # Extract address (العنوان)
                try:
                    elem = self.driver.find_element(By.XPATH, "//*[contains(text(), 'العنوان')]")
                    parent_text = elem.find_element(By.XPATH, "./..").text
                    if 'العنوان' in parent_text:
                        # Handle both "العنوان:" and "العنوان :" formats
                        match = re.search(r'العنوان\s*[:\s]+([^\n]+)', parent_text)
                        if match:
                            data['address'] = match.group(1).strip()
                except:
                    match = re.search(r'العنوان\s*[:\s]+([^\n]+)', full_text)
                    if match:
                        data['address'] = match.group(1).strip()
                
                # Extract subcommittee number (رقم اللجنة الفرعية)
                try:
                    # Try multiple XPath patterns
                    patterns = [
                        "//*[contains(text(), 'رقم اللجنة الفرعية')]",
                        "//*[contains(., 'رقم اللجنة الفرعية')]",
                        "//text()[contains(., 'رقم اللجنة الفرعية')]/.."
                    ]
                    found = False
                    for pattern in patterns:
                        try:
                            elem = self.driver.find_element(By.XPATH, pattern)
                            # Get the text from parent or the element itself
                            try:
                                parent = elem.find_element(By.XPATH, "./..")
                                text = parent.text
                            except:
                                text = elem.text
                            
                            # Extract number using regex
                            match = re.search(r'رقم اللجنة الفرعية\s*[:\s]+\s*([^\n]+)', text)
                            if match:
                                num_str = match.group(1).strip()
                                num_str = self.convert_arabic_numerals(num_str)
                                num = re.search(r'\d+', num_str)
                                if num:
                                    data['subcommittee_number'] = num.group()
                                    found = True
                                    break
                        except:
                            continue
                    
                    if not found:
                        # Fallback to full text regex
                        match = re.search(r'رقم اللجنة الفرعية\s*[:\s]+\s*([^\n]+)', full_text)
                        if match:
                            num_str = match.group(1).strip()
                            num_str = self.convert_arabic_numerals(num_str)
                            num = re.search(r'\d+', num_str)
                            if num:
                                data['subcommittee_number'] = num.group()
                except Exception as e:
                    logger.debug(f"Error extracting subcommittee number: {e}")
                    # Try regex on full text
                    match = re.search(r'رقم اللجنة الفرعية\s*[:\s]+\s*([^\n]+)', full_text)
                    if match:
                        num_str = match.group(1).strip()
                        num_str = self.convert_arabic_numerals(num_str)
                        num = re.search(r'\d+', num_str)
                        if num:
                            data['subcommittee_number'] = num.group()
                
                # Extract electoral list number (رقمك في الكشوف الانتخابية)
                try:
                    # Try multiple XPath patterns
                    patterns = [
                        "//*[contains(text(), 'رقمك في الكشوف الانتخابية')]",
                        "//*[contains(., 'رقمك في الكشوف الانتخابية')]",
                        "//text()[contains(., 'رقمك في الكشوف الانتخابية')]/.."
                    ]
                    found = False
                    for pattern in patterns:
                        try:
                            elem = self.driver.find_element(By.XPATH, pattern)
                            # Get the text from parent or the element itself
                            try:
                                parent = elem.find_element(By.XPATH, "./..")
                                text = parent.text
                            except:
                                text = elem.text
                            
                            # Extract number using regex
                            match = re.search(r'رقمك في الكشوف الانتخابية\s*[:\s]+\s*([^\n]+)', text)
                            if match:
                                num_str = match.group(1).strip()
                                num_str = self.convert_arabic_numerals(num_str)
                                num = re.search(r'\d+', num_str)
                                if num:
                                    data['electoral_list_number'] = num.group()
                                    found = True
                                    break
                        except:
                            continue
                    
                    if not found:
                        # Fallback to full text regex
                        match = re.search(r'رقمك في الكشوف الانتخابية\s*[:\s]+\s*([^\n]+)', full_text)
                        if match:
                            num_str = match.group(1).strip()
                            num_str = self.convert_arabic_numerals(num_str)
                            num = re.search(r'\d+', num_str)
                            if num:
                                data['electoral_list_number'] = num.group()
                except Exception as e:
                    logger.debug(f"Error extracting electoral list number: {e}")
                    # Try regex on full text
                    match = re.search(r'رقمك في الكشوف الانتخابية\s*[:\s]+\s*([^\n]+)', full_text)
                    if match:
                        num_str = match.group(1).strip()
                        num_str = self.convert_arabic_numerals(num_str)
                        num = re.search(r'\d+', num_str)
                        if num:
                            data['electoral_list_number'] = num.group()
                
            except Exception as e:
                logger.debug(f"XPath extraction failed, trying regex: {e}")
                # Fallback to regex-only extraction
                # Extract electoral center (مركزك الإنتخابي)
                match = re.search(r'مركزك الإنتخابي[:\s]+([^\n]+)', full_text)
                if match:
                    data['electoral_center'] = match.group(1).strip()
                
                # Extract district (قسم)
                match = re.search(r'قسم[:\s]+([^\n]+)', full_text)
                if match:
                    data['district'] = match.group(1).strip()
                
                # Extract address (العنوان)
                match = re.search(r'العنوان\s*[:\s]+([^\n]+)', full_text)
                if match:
                    data['address'] = match.group(1).strip()
                
                # Extract subcommittee number (رقم اللجنة الفرعية)
                match = re.search(r'رقم اللجنة الفرعية[:\s]+([^\n]+)', full_text)
                if match:
                    num_str = match.group(1).strip()
                    num_str = self.convert_arabic_numerals(num_str)
                    num = re.search(r'\d+', num_str)
                    if num:
                        data['subcommittee_number'] = num.group()
                
                # Extract electoral list number (رقمك في الكشوف الانتخابية)
                match = re.search(r'رقمك في الكشوف الانتخابية[:\s]+([^\n]+)', full_text)
                if match:
                    num_str = match.group(1).strip()
                    num_str = self.convert_arabic_numerals(num_str)
                    num = re.search(r'\d+', num_str)
                    if num:
                        data['electoral_list_number'] = num.group()
            
            logger.info(f"Extracted data: {data}")
            
            return {
                'success': True,
                'data': data
            }
                
        except Exception as e:
            logger.error(f"Error in extract_result_data: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def convert_arabic_numerals(self, text: str) -> str:
        """Convert Arabic numerals to English numerals."""
        arabic_to_english = {
            '٠': '0', '١': '1', '٢': '2', '٣': '3', '٤': '4',
            '٥': '5', '٦': '6', '٧': '7', '٨': '8', '٩': '9'
        }
        result = text
        for arabic, english in arabic_to_english.items():
            result = result.replace(arabic, english)
        return result
    
    def close(self):
        """Close the browser."""
        if self.driver:
            self.driver.quit()
            logger.info("Browser closed")


def convert_csv_to_excel(csv_file: str, excel_file: str = None) -> str:
    """Convert CSV file to Excel format."""
    import pandas as pd
    
    if excel_file is None:
        excel_file = csv_file.replace('.csv', '.xlsx')
    
    logger.info(f"Converting CSV to Excel: {csv_file} -> {excel_file}")
    
    # Try different encodings
    encodings = ['utf-8-sig', 'utf-8', 'latin-1']
    df = None
    
    for encoding in encodings:
        try:
            df = pd.read_csv(csv_file, encoding=encoding)
            logger.info(f"Successfully converted using {encoding} encoding")
            break
        except UnicodeDecodeError:
            continue
    
    if df is None:
        raise ValueError(f"Could not read CSV file with any encoding: {csv_file}")
    
    df.to_excel(excel_file, index=False)
    return excel_file


def process_excel_file(input_file: str, output_file: str = None, max_rows: int = None, headless: bool = False):
    """
    Process Excel file: read national IDs, scrape website, write results.
    
    Args:
        input_file: Path to input Excel file
        output_file: Path to output Excel file (default: input_file with _results suffix)
        max_rows: Maximum number of rows to process (None = all rows)
        headless: Run browser in headless mode
    """
    # Handle CSV files
    if input_file.lower().endswith('.csv'):
        logger.info("CSV file detected, converting to Excel format...")
        excel_file = convert_csv_to_excel(input_file)
        logger.info(f"Using temporary Excel file: {excel_file}")
        input_file = excel_file
    
    # Load Excel file
    logger.info(f"Loading Excel file: {input_file}")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    # Find national ID column
    national_id_col = None
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        if cell_value and 'الرقم القومي' in str(cell_value):
            national_id_col = col_idx
            break
    
    if national_id_col is None:
        raise ValueError("Could not find 'الرقم القومي' column in Excel file")
    
    logger.info(f"Found national ID column: '{ws.cell(row=1, column=national_id_col).value}' at column {national_id_col}")
    
    # Define output columns mapping
    output_columns = {
        'المركز الانتخابي': None,
        'الدائرة': None,
        'العنوان': None,
        'رقم اللجنة الفرعية': None,
        'رقمك في الكشوف الانتخابية': None,
        'الحالة': None
    }
    
    # Find or create output columns
    for col_name in output_columns.keys():
        found = False
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            if cell_value and col_name in str(cell_value):
                output_columns[col_name] = col_idx
                logger.info(f"Found existing column: '{col_name}' at column {col_idx}")
                found = True
                break
        
        if not found:
            # Create new column
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value=col_name)
            output_columns[col_name] = new_col
            logger.info(f"Created new column: '{col_name}' at column {new_col}")
    
    # Initialize scraper
    scraper = FreeElectionsScraper(headless=headless)
    
    # Process rows
    total_rows = ws.max_row
    processed = 0
    skipped = 0
    errors = 0
    
    if max_rows:
        end_row = min(2 + max_rows, total_rows + 1)
        logger.info(f"Processing {max_rows} row(s) for testing")
    else:
        end_row = total_rows + 1
    
    try:
        for row_idx in range(2, end_row):
            # Get national ID
            national_id_cell = ws.cell(row=row_idx, column=national_id_col)
            national_id = national_id_cell.value
            
            # Skip empty cells
            if national_id is None or str(national_id).strip() == '':
                skipped += 1
                continue
            
            national_id = str(national_id).strip()
            logger.info(f"Processing row {row_idx}/{total_rows}: National ID = {national_id}")
            
            # Scrape data
            result = scraper.scrape_electoral_data(national_id)
            
            if result['success']:
                data = result['data']
                status = data.get('status', 'success')
                
                # Handle different status cases
                if status == 'underage':
                    logger.info(f"National ID {national_id}: Underage - not eligible to vote")
                    ws.cell(row=row_idx, column=output_columns['الحالة'], 
                           value='ليس له حق الانتخاب - تحت السن')
                    processed += 1
                    
                elif status == 'not_registered':
                    logger.info(f"National ID {national_id}: Not registered in voter database")
                    ws.cell(row=row_idx, column=output_columns['الحالة'], 
                           value='ليس له حق الانتخاب')
                    processed += 1
                    
                elif status == 'success':
                    district = data.get('district', '')
                    
                    # Only fill data if district is in allowed list
                    if district in ALLOWED_DISTRICTS:
                        logger.info(f"District '{district}' is in allowed list. Filling data...")
                        
                        # Fill the data
                        ws.cell(row=row_idx, column=output_columns['المركز الانتخابي'], 
                               value=data.get('electoral_center', ''))
                        ws.cell(row=row_idx, column=output_columns['الدائرة'], 
                               value=data.get('district', ''))
                        ws.cell(row=row_idx, column=output_columns['العنوان'], 
                               value=data.get('address', ''))
                        ws.cell(row=row_idx, column=output_columns['رقم اللجنة الفرعية'], 
                               value=data.get('subcommittee_number', ''))
                        ws.cell(row=row_idx, column=output_columns['رقمك في الكشوف الانتخابية'], 
                               value=data.get('electoral_list_number', ''))
                        ws.cell(row=row_idx, column=output_columns['الحالة'], 
                               value='له حق الانتخاب')
                        
                        processed += 1
                    else:
                        logger.info(f"District '{district}' is not in allowed list. Skipping...")
                        skipped += 1
                else:
                    logger.warning(f"Unknown status: {status}")
                    errors += 1
            else:
                logger.error(f"Failed to scrape data for {national_id}: {result.get('error', 'Unknown error')}")
                errors += 1
            
            # Small delay between requests
            time.sleep(1)
    
    finally:
        # Close browser
        scraper.close()
        
        # Save results
        if output_file is None:
            base_name = Path(input_file).stem
            output_file = f"{base_name}_results.xlsx"
        
        logger.info(f"Saved results to: {output_file}")
        wb.save(output_file)
        logger.info(f"Summary: {processed} processed, {skipped} skipped, {errors} errors")
        logger.info("Processing complete!")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='FREE Electoral Data Scraper - No AI/API costs!',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process first 5 rows for testing
  python selenium_scraper.py spreadsheet.xlsx --limit 5
  
  # Process all rows
  python selenium_scraper.py spreadsheet.xlsx
  
  # Process CSV file (will be converted to Excel)
  python selenium_scraper.py input.csv
  
  # Run in headless mode (no browser window)
  python selenium_scraper.py spreadsheet.xlsx --headless
        """
    )
    
    parser.add_argument(
        'input_file',
        help='Path to input Excel or CSV file'
    )
    parser.add_argument(
        '-o', '--output',
        default=None,
        help='Path to output Excel file (default: input_file with _results suffix)'
    )
    parser.add_argument(
        '--limit',
        type=int,
        default=None,
        help='Limit number of rows to process (useful for testing)'
    )
    parser.add_argument(
        '--headless',
        action='store_true',
        help='Run browser in headless mode (no GUI)'
    )
    
    args = parser.parse_args()
    
    # Check if input file exists
    if not Path(args.input_file).exists():
        logger.error(f"Input file not found: {args.input_file}")
        sys.exit(1)
    
    # Process the file
    try:
        process_excel_file(
            args.input_file,
            args.output,
            args.limit,
            args.headless
        )
    except KeyboardInterrupt:
        logger.info("Process interrupted by user")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

